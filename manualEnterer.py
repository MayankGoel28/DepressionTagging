import pandas as pd
import openpyxl
legaltags = ('health', 'finance', 'family loss', 'relationships',
             'personal', 'pets', 'assault', 'harassment', 'anxiety', 'hl', 'fn', 'fl', 'rl', 'pr', 'pt', 'as', 'hr', 'an')
fulltags = ('health', 'finance', 'family loss', 'relationships',
            'personal', 'pets', 'assault', 'harassment', 'anxiety')
abtags = ('hl', 'fn', 'fl', 'rl', 'pr', 'pt', 'as', 'hr', 'an')


def callHelp():
    print(legaltags)


def checkIf(tag):
    if tag == 'NaN':
        return False
    flag = 0
    for checkForTag in legaltags:
        checkForTag = str(checkForTag).lower().strip()
        tag = str(tag).lower().strip()
        if checkForTag == tag:
            flag = 1
    if flag == 0:
        return False
    return True


data = pd.read_excel(
    r'./analyse.xlsx')
srcfile = openpyxl.load_workbook(
    './analyse.xlsx', read_only=False, keep_vba=True)
sheet = srcfile.active
tags = pd.DataFrame(data, columns=['Tags'])
taglist = list(tags['Tags'])
c = 2
gender = ""
answer = ""
clearline = "\n"*100
for x in taglist:
    if x in abtags:
        indexOftag = abtags.index(x)
        sheet.cell(row=c, column=3).value = fulltags[indexOftag]
        c+=1
c = 2
for x in taglist:
    if not checkIf(x):
        print(clearline)
        print(sheet.cell(row=c, column=1).value)
        print("Enter Gender.")
        gender = input()
        gender = gender.strip()
        if gender == 'help':
            callHelp()
            gender = input()
        if gender == 'exit':
            break
        print("Enter tag.")
        answer = input()
        answer = answer.strip()
        sheet.cell(row=c, column=2).value = gender
        sheet.cell(row=c, column=3).value = answer
    c += 1
srcfile.save('./analyse.xlsx')
